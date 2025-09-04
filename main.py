import pandas as pd
import io
import functions_framework
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment

# ==============================================================================
# LÓGICA DE NEGOCIO (Tus funciones para crear reportes)
# Estas funciones permanecen casi idénticas, ya que su lógica interna no
# dependía directamente de Flask.
# ==============================================================================

def crear_reporte_psicologos(datos_requisiciones):
    """
    TRANSFORMADO: Genera un reporte analítico que incluye una hoja de datos con rangos de tiempo,
    un resumen por analista y un dashboard con tablas y gráficos.
    Las hojas de Analista y Empresas están ordenadas por total y tienen subtotales por grupo.
    """
    # --- 1. Preparación y Limpieza de Datos ---
    df = pd.DataFrame(datos_requisiciones)
    
    columna_estado = 'Estado'
    if columna_estado in df.columns:
        df = df[df[columna_estado] != 'Cerrado'].copy()
    
    columnas_a_eliminar = ['Adicionales', 'Contratados a tiempo', 'No contratados a tiempo']
    df.drop(columns=columnas_a_eliminar, inplace=True, errors='ignore')

    dias_col_original = None
    if 'Días' in df.columns: dias_col_original = 'Días'
    elif 'Dias' in df.columns: dias_col_original = 'Dias'
    elif 'dias' in df.columns: dias_col_original = 'dias'
    
    if dias_col_original is None:
        raise KeyError(f"No se encontró la columna de días ('Días' o 'Dias'). Columnas recibidas: {list(df.columns)}")

    df.rename(columns={dias_col_original: 'Dias', 'Psicólogo': 'Psicologo', 'Compañía': 'Compania'}, inplace=True)
    
    df["Dias"] = pd.to_numeric(df["Dias"], errors="coerce").fillna(0).astype(int)
    df["Pendientes"] = pd.to_numeric(df["Pendientes"], errors="coerce").fillna(0).astype(int)

    def categorizar_dias(dias):
        if 0 <= dias <= 20: return "Entre 0 y 20"
        elif 21 <= dias <= 50: return "Entre 21 y 50"
        else: return "Mayor a 50"

    df["Tiempo"] = df["Dias"].apply(categorizar_dias)
    
    # --- 2. Creación de las Tablas de Resumen ---
    df_analista = df.groupby(['Psicologo', 'Compania'])['Pendientes'].sum().reset_index()
    df_analista.rename(columns={'Pendientes': 'Total', 'Psicologo': 'Psicólogo', 'Compania': 'Compañía'}, inplace=True)

    df_detalle_empresa_tipo = df.groupby(['Compania', 'Tipo'])['Pendientes'].sum().reset_index()
    df_detalle_empresa_tipo.rename(columns={'Pendientes': 'Total', 'Compania': 'Compañía'}, inplace=True)
    
    df_resumen_tipo = df.groupby('Tipo')['Pendientes'].sum().reset_index()
    df_resumen_tipo.rename(columns={'Pendientes': 'Total'}, inplace=True)

    df_resumen_tiempo = df.groupby('Tiempo')['Pendientes'].sum().reset_index()
    df_resumen_tiempo.rename(columns={'Pendientes': 'Total'}, inplace=True)

    se_anadio_fila_fantasma = False
    if len(df_resumen_tiempo) == 1:
        fila_fantasma = pd.DataFrame([{'Tiempo': '(Otra categoría sin datos)', 'Total': 0}])
        df_resumen_tiempo = pd.concat([df_resumen_tiempo, fila_fantasma], ignore_index=True)
        se_anadio_fila_fantasma = True

    if not df_resumen_tipo.empty:
        total_tipo = df_resumen_tipo['Total'].sum()
        fila_total_tipo = pd.DataFrame([{'Tipo': 'Total', 'Total': total_tipo}])
        df_resumen_tipo = pd.concat([df_resumen_tipo, fila_total_tipo], ignore_index=True)

    if not df_resumen_tiempo.empty:
        total_tiempo = df_resumen_tiempo['Total'].sum()
        fila_total_tiempo = pd.DataFrame([{'Tiempo': 'Total', 'Total': total_tiempo}])
        df_resumen_tiempo = pd.concat([df_resumen_tiempo, fila_total_tiempo], ignore_index=True)

    # --- 3. Escritura del Archivo Excel con Formato y Gráficos ---
    output_buffer = io.BytesIO()
    with pd.ExcelWriter(output_buffer, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Requisiciones', index=False)
        workbook = writer.book
        header_format = workbook.add_format({'bold': True, 'bg_color': '#DDEBF7', 'border': 1})
        total_format = workbook.add_format({'bold': True, 'bg_color': '#F2F2F2'})
        
        # Hoja de Analista
        worksheet_analista = workbook.add_worksheet('Analista')
        analyst_totals = df_analista.groupby('Psicólogo')['Total'].sum().sort_values(ascending=False)
        worksheet_analista.write_row('A1', ['Psicólogo', 'Compañía', 'Suma de Pendientes'], header_format)
        current_row_analista = 1
        for psicologo in analyst_totals.index:
            df_filtrado = df_analista[df_analista['Psicólogo'] == psicologo].sort_values(by='Total', ascending=False)
            for _, row_data in df_filtrado.iterrows():
                worksheet_analista.write_row(current_row_analista, 0, [row_data['Psicólogo'], row_data['Compañía'], row_data['Total']])
                current_row_analista += 1
            total_psicologo = analyst_totals[psicologo]
            worksheet_analista.write(current_row_analista, 0, f"Total {psicologo}", total_format)
            worksheet_analista.write(current_row_analista, 2, total_psicologo, total_format)
            current_row_analista += 1
        worksheet_analista.write(current_row_analista, 0, "Total General", header_format)
        worksheet_analista.write(current_row_analista, 2, df_analista['Total'].sum(), header_format)
        
        # Hoja de Empresas
        worksheet_resumen = workbook.add_worksheet('Empresas')
        empresa_totals = df_detalle_empresa_tipo.groupby('Compañía')['Total'].sum().sort_values(ascending=False)
        worksheet_resumen.write_row('A1', ['Compañía', 'Tipo', 'Total'], header_format)
        current_row_resumen = 1
        for compania in empresa_totals.index:
            df_filtrado_compania = df_detalle_empresa_tipo[df_detalle_empresa_tipo['Compañía'] == compania]
            for _, row_data in df_filtrado_compania.iterrows():
                worksheet_resumen.write_row(current_row_resumen, 0, [row_data['Compañía'], row_data['Tipo'], row_data['Total']])
                current_row_resumen += 1
            total_compania = empresa_totals[compania]
            worksheet_resumen.write(current_row_resumen, 0, f"Total {compania}", total_format)
            worksheet_resumen.write(current_row_resumen, 2, total_compania, total_format)
            current_row_resumen += 1
        worksheet_resumen.write(current_row_resumen, 0, "Total General", header_format)
        worksheet_resumen.write(current_row_resumen, 2, df_detalle_empresa_tipo['Total'].sum(), header_format)
        
        # Tablas de resumen y gráficos
        df_resumen_tipo.to_excel(writer, sheet_name='Empresas', startrow=1, startcol=4, header=['Tipo', 'Total'], index=False)
        start_row_tiempo_table = len(df_resumen_tipo) + 4
        df_resumen_tiempo.to_excel(writer, sheet_name='Empresas', startrow=start_row_tiempo_table, startcol=4, header=['Tiempo', 'Total'], index=False)

        chart_tipo = workbook.add_chart({'type': 'doughnut'})
        num_categorias_tipo = len(df_resumen_tipo) - 1
        end_row_chart_tipo = start_row_chart_tipo + num_categorias_tipo - 1
        cat_range_tipo = f"'Empresas'!$E${start_row_chart_tipo}:$E${end_row_chart_tipo}"
        val_range_tipo = f"'Empresas'!$F${start_row_chart_tipo}:$F${end_row_chart_tipo}"
        
        chart_tipo.add_series({
            'name': 'Resumen por Tipo',
            'categories': cat_range_tipo,
            'values': val_range_tipo,
            'data_labels': {'percentage': True, 'leader_lines': True, 'num_format': '0%;;'}
        })
        chart_tipo.set_title({'name': 'Tipo de proceso'})
        worksheet_resumen.insert_chart('H2', chart_tipo)
        chart_tiempo = workbook.add_chart({'type': 'doughnut'})
        start_row_tiempo_table = len(df_resumen_tipo) + 4
        num_categorias_tiempo = len(df_resumen_tiempo) - 1
        start_row_chart_tiempo = start_row_tiempo_table + 2
        end_row_chart_tiempo = start_row_chart_tiempo + num_categorias_tiempo - 1
        cat_range_tiempo = f"'Empresas'!$E${start_row_chart_tiempo}:$E${end_row_chart_tiempo}"
        val_range_tiempo = f"'Empresas'!$F${start_row_chart_tiempo}:$F${end_row_chart_tiempo}"
        
        chart_tiempo.add_series({
            'name': 'Resumen por Tiempo',
            'categories': cat_range_tiempo,
            'values': val_range_tiempo,
            'data_labels': {'percentage': True, 'leader_lines': True, 'num_format': '0%;;'}
        })
        chart_tiempo.set_title({'name': 'Tiempo Transcurrido'})
        worksheet_resumen.insert_chart('H15', chart_tiempo)
        worksheet_datos = writer.sheets['Requisiciones']
        for idx, col in enumerate(df.columns):
            max_len = max(df[col].astype(str).map(len).max(), len(col)) + 3
            worksheet_datos.set_column(idx, idx, max_len)
        worksheet_analista.set_column('A:A', 30); worksheet_analista.set_column('B:B', 45); worksheet_analista.set_column('C:C', 20)
        worksheet_resumen.set_column('A:A', 45); worksheet_resumen.set_column('B:B', 25); worksheet_resumen.set_column('C:C', 15); worksheet_resumen.set_column('E:E', 25); worksheet_resumen.set_column('F:F', 15)

    output_buffer.seek(0)
    return output_buffer

# --- Lógica para el reporte "General" ---
def crear_reporte_completo(datos_resumen_psicologo, datos_requisiciones):
    NOMBRE_HOJA_DATOS = "Requisiciones"
    NOMBRE_HOJA_PSICOLOGO = "Resumen psicólogo"
    COLUMNA_TIPO = "Tipo"
    COLUMNA_ESTADO = "Estado"
    nombre_hoja_resumen = "Resumen Dinamico"
    
    df_psicologo = pd.DataFrame(datos_resumen_psicologo)
    df_requisiciones = pd.DataFrame(datos_requisiciones)

    categorias_tipo = df_requisiciones[COLUMNA_TIPO].dropna().unique()
    categorias_estado = df_requisiciones[COLUMNA_ESTADO].dropna().unique()

    output_buffer = io.BytesIO()
    with pd.ExcelWriter(output_buffer, engine="openpyxl") as writer:
        df_psicologo.to_excel(writer, sheet_name=NOMBRE_HOJA_PSICOLOGO, index=False)
        df_requisiciones.to_excel(writer, sheet_name=NOMBRE_HOJA_DATOS, index=False)
        book = writer.book
        ws_resumen = book.create_sheet(nombre_hoja_resumen)

        def get_col_letter_by_name(dataframe, col_name):
            try: return get_column_letter(dataframe.columns.get_loc(col_name) + 1)
            except KeyError: return None

        col_letra_tipo = get_col_letter_by_name(df_requisiciones, COLUMNA_TIPO)
        col_letra_estado = get_col_letter_by_name(df_requisiciones, COLUMNA_ESTADO)
        fila_actual = 2

        if col_letra_tipo:
            ws_resumen.cell(row=fila_actual, column=1, value=COLUMNA_TIPO)
            ws_resumen.cell(row=fila_actual, column=2, value="Suma de Solicitados")
            fila_inicio_tipo = fila_actual + 1
            for cat in categorias_tipo:
                fila_actual += 1
                ws_resumen.cell(row=fila_actual, column=1, value=cat)
                ws_resumen.cell(row=fila_actual, column=2, value=f'=COUNTIF({NOMBRE_HOJA_DATOS}!{col_letra_tipo}:{col_letra_tipo}, "{cat}")')
            fila_actual += 1
            ws_resumen.cell(row=fila_actual, column=1, value="Suma total")
            ws_resumen.cell(row=fila_actual, column=2, value=f'=SUM(B{fila_inicio_tipo}:B{fila_actual-1})')
            fila_actual += 3

        if col_letra_estado:
            ws_resumen.cell(row=fila_actual, column=1, value=COLUMNA_ESTADO)
            ws_resumen.cell(row=fila_actual, column=2, value="Suma de Solicitados")
            fila_inicio_estado = fila_actual + 1
            for cat in categorias_estado:
                fila_actual += 1
                ws_resumen.cell(row=fila_actual, column=1, value=cat)
                ws_resumen.cell(row=fila_actual, column=2, value=f'=COUNTIF({NOMBRE_HOJA_DATOS}!{col_letra_estado}:{col_letra_estado}, "{cat}")')
            fila_actual += 1
            ws_resumen.cell(row=fila_actual, column=1, value="Suma total")
            ws_resumen.cell(row=fila_actual, column=2, value=f'=SUM(B{fila_inicio_estado}:B{fila_actual-1})')

        # Auto-ajuste de columnas para todas las hojas
        sheets_data = { NOMBRE_HOJA_PSICOLOGO: df_psicologo, NOMBRE_HOJA_DATOS: df_requisiciones }
        for sheet_name, df in sheets_data.items():
            worksheet = writer.sheets[sheet_name]
            for idx, col in enumerate(df.columns, 1):
                max_len = max(df[col].astype(str).map(len).max(), len(str(col))) + 2
                worksheet.column_dimensions[get_column_letter(idx)].width = max_len
        ws_resumen.column_dimensions['A'].width = 25
        ws_resumen.column_dimensions['B'].width = 25

    output_buffer.seek(0)
    return output_buffer

# --- Lógica para el reporte "Formateado" ---
def crear_reporte_formateado_con_resumen(datos_psicologo, datos_requisiciones):
    NOMBRE_HOJA_RESUMEN = "Resumen Psicólogo"
    NOMBRE_HOJA_REQUISICIONES = "Requisiciones"
    PALABRAS_POR_LINEA_JUSTIFICACION = 20
    ANCHO_FIJO_COLUMNA_JUSTIFICACION = 60

    def limpiar_justificacion(texto):
        if not isinstance(texto, str): return texto
        pos = texto.rfind(':')
        return texto[pos + 1:].strip() if pos != -1 and '-' in texto[:pos] else texto

    def ajustar_texto_por_palabras(texto, max_palabras):
        if not isinstance(texto, str): return texto
        palabras = texto.split(' ')
        if len(palabras) <= max_palabras: return texto
        return '\n'.join([' '.join(palabras[i:i + max_palabras]) for i in range(0, len(palabras), max_palabras)])

    df_psicologo = pd.DataFrame(datos_psicologo)
    df_requisiciones = pd.DataFrame(datos_requisiciones)
    
    if 'Justificacion' in df_requisiciones.columns:
        df_requisiciones.rename(columns={'Justificacion': 'Justificación'}, inplace=True)

    if 'Justificación' in df_requisiciones.columns:
        df_requisiciones['Justificación'] = df_requisiciones['Justificación'].apply(limpiar_justificacion).apply(
            lambda x: ajustar_texto_por_palabras(x, PALABRAS_POR_LINEA_JUSTIFICACION))

    output_buffer = io.BytesIO()
    with pd.ExcelWriter(output_buffer, engine="openpyxl") as writer:
        df_psicologo.to_excel(writer, sheet_name=NOMBRE_HOJA_RESUMEN, index=False)
        df_requisiciones.to_excel(writer, sheet_name=NOMBRE_HOJA_REQUISICIONES, index=False)

        # Formato Hoja "Resumen Psicólogo"
        ws_resumen = writer.sheets[NOMBRE_HOJA_RESUMEN]
        for idx, col_name in enumerate(df_psicologo.columns, 1):
            try:
                max_len = max(df_psicologo[col_name].astype(str).map(len).max(), len(str(col_name))) + 2
                ws_resumen.column_dimensions[get_column_letter(idx)].width = max_len
            except (ValueError, KeyError):
                ws_resumen.column_dimensions[get_column_letter(idx)].width = len(str(col_name)) + 5

        # Formato Hoja "Requisiciones"
        ws_req = writer.sheets[NOMBRE_HOJA_REQUISICIONES]
        for idx, col_name in enumerate(df_requisiciones.columns, 1):
            col_letter = get_column_letter(idx)
            if col_name == 'Justificación':
                ws_req.column_dimensions[col_letter].width = ANCHO_FIJO_COLUMNA_JUSTIFICACION
                for fila in range(2, ws_req.max_row + 1):
                    cell = ws_req.cell(row=fila, column=idx)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            else:
                try:
                    max_len = max(df_requisiciones[col_name].astype(str).map(len).max(), len(str(col_name))) + 2
                    ws_req.column_dimensions[col_letter].width = max_len
                except (ValueError, KeyError):
                    ws_req.column_dimensions[col_letter].width = len(str(col_name)) + 5

    output_buffer.seek(0)
    return output_buffer


# ==============================================================================
# FUNCIÓN DE ENTRADA PRINCIPAL PARA GOOGLE CLOUD FUNCTIONS
# ==============================================================================

@functions_framework.http
def excel_report_handler(request):
    """
    Función HTTP de Cloud Functions que actúa como enrutador.
    Invoca la lógica de generación de reportes correcta según la ruta de la URL.
    """
    # El objeto 'request' es proporcionado por el entorno de Cloud Functions.
    # Es similar al objeto 'request' de Flask.
    
    # Obtenemos la ruta de la solicitud para decidir qué hacer.
    # Ej: '/generar-reporte-psicologos'
    request_path = request.path

    # Obtenemos los datos JSON del cuerpo de la solicitud.
    try:
        datos_json = request.get_json(silent=True)
        if not datos_json:
            return ("No se recibieron datos JSON válidos.", 400)
    except Exception as e:
        return (f"Error al parsear el JSON: {repr(e)}", 400)

    try:
        buffer_excel = None
        download_name = "reporte.xlsx"

        # --- Enrutador basado en la ruta ---
        if request_path.endswith('/generar-reporte-psicologos'):
            lista_de_requisiciones = datos_json.get('datos_requisiciones')
            if lista_de_requisiciones is None:
                return ("El JSON debe contener la clave 'datos_requisiciones'.", 400)
            
            buffer_excel = crear_reporte_psicologos(lista_de_requisiciones)
            download_name = "reporte_analitico_psicologos.xlsx"

        elif request_path.endswith('/generar-resumen-dinamico'):
            datos_requisiciones = datos_json.get('datos_requisiciones')
            datos_psicologo = datos_json.get('datos_resumen_psicologo')
            if datos_requisiciones is None or datos_psicologo is None:
                return ("El JSON debe contener las claves 'datos_requisiciones' y 'datos_resumen_psicologo'.", 400)
                
            buffer_excel = crear_reporte_completo(datos_psicologo, datos_requisiciones)
            download_name = "Reporte_General_Completo.xlsx"

        elif request_path.endswith('/formatear-reporte-general'):
            datos_requisiciones = datos_json.get('datos_requisiciones')
            datos_psicologo = datos_json.get('datos_resumen_psicologo')
            if datos_requisiciones is None or datos_psicologo is None:
                return ("El JSON debe contener las claves 'datos_requisiciones' y 'datos_resumen_psicologo'.", 400)

            buffer_excel = crear_reporte_formateado_con_resumen(datos_psicologo, datos_requisiciones)
            download_name = "Reporte_General_Formateado.xlsx"

        else:
            return ("Ruta no reconocida. Use una de las rutas de generación de reportes válidas.", 404)

        # Si se generó un archivo, lo retornamos
        if buffer_excel:
            # Creamos la respuesta manualmente, similar a send_file de Flask.
            headers = {
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'Content-Disposition': f'attachment;filename="{download_name}"'
            }
            return (buffer_excel.getvalue(), 200, headers)
        else:
            # Esto no debería ocurrir si la ruta es correcta, pero es una salvaguarda.
            return ("No se pudo generar el archivo para la ruta especificada.", 500)

    except Exception as e:
        # Capturamos cualquier error inesperado durante la generación del Excel.
        import traceback
        traceback.print_exc()
        return (f"Ocurrió un error interno en el servidor: {repr(e)}", 500)